#!/usr/bin/env python3
"""
Personal Finance Dashboard Backend API Tests
Tests all backend endpoints for the finance dashboard application.
"""

import requests
import json
import os
from datetime import datetime, timedelta
import uuid

# Get base URL from environment
BASE_URL = os.getenv('NEXT_PUBLIC_BASE_URL', 'https://finview-13.preview.emergentagent.com')
API_BASE = f"{BASE_URL}/api"

class FinanceAPITester:
    def __init__(self):
        self.session = requests.Session()
        self.test_results = []
        self.created_resources = {
            'accounts': [],
            'categories': [],
            'transactions': []
        }
    
    def log_result(self, test_name, success, message, details=None):
        """Log test result"""
        result = {
            'test': test_name,
            'success': success,
            'message': message,
            'details': details or {}
        }
        self.test_results.append(result)
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
        if details and not success:
            print(f"   Details: {details}")
    
    def test_get_accounts(self):
        """Test GET /api/accounts endpoint"""
        try:
            response = self.session.get(f"{API_BASE}/accounts")
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    accounts = data['data']
                    self.log_result(
                        "GET /api/accounts", 
                        True, 
                        f"Retrieved {len(accounts)} accounts successfully",
                        {'account_count': len(accounts), 'sample_account': accounts[0] if accounts else None}
                    )
                    return accounts
                else:
                    self.log_result("GET /api/accounts", False, "Invalid response format", data)
            else:
                self.log_result("GET /api/accounts", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("GET /api/accounts", False, f"Request failed: {str(e)}")
        return []
    
    def test_create_account(self):
        """Test POST /api/accounts endpoint"""
        try:
            test_account = {
                "name": f"Test Savings Account {uuid.uuid4().hex[:8]}",
                "type": "BANK",
                "balance": 1500.50
            }
            
            response = self.session.post(
                f"{API_BASE}/accounts",
                json=test_account,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    account = data['data']
                    self.created_resources['accounts'].append(account['id'])
                    self.log_result(
                        "POST /api/accounts", 
                        True, 
                        f"Created account '{account['name']}'",
                        {'account_id': account['id'], 'balance': account['balance']}
                    )
                    return account
                else:
                    self.log_result("POST /api/accounts", False, "Invalid response format", data)
            else:
                self.log_result("POST /api/accounts", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("POST /api/accounts", False, f"Request failed: {str(e)}")
        return None
    
    def test_get_categories(self):
        """Test GET /api/categories endpoint"""
        try:
            response = self.session.get(f"{API_BASE}/categories")
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    categories = data['data']
                    self.log_result(
                        "GET /api/categories", 
                        True, 
                        f"Retrieved {len(categories)} categories successfully",
                        {'category_count': len(categories), 'sample_category': categories[0] if categories else None}
                    )
                    return categories
                else:
                    self.log_result("GET /api/categories", False, "Invalid response format", data)
            else:
                self.log_result("GET /api/categories", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("GET /api/categories", False, f"Request failed: {str(e)}")
        return []
    
    def test_create_category(self):
        """Test POST /api/categories endpoint"""
        try:
            test_category = {
                "name": f"Test Entertainment {uuid.uuid4().hex[:8]}",
                "type": "EXPENSE",
                "color": "#ff6b6b"
            }
            
            response = self.session.post(
                f"{API_BASE}/categories",
                json=test_category,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    category = data['data']
                    self.created_resources['categories'].append(category['id'])
                    self.log_result(
                        "POST /api/categories", 
                        True, 
                        f"Created category '{category['name']}'",
                        {'category_id': category['id'], 'type': category['type']}
                    )
                    return category
                else:
                    self.log_result("POST /api/categories", False, "Invalid response format", data)
            else:
                self.log_result("POST /api/categories", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("POST /api/categories", False, f"Request failed: {str(e)}")
        return None
    
    def test_get_transactions(self):
        """Test GET /api/transactions endpoint with various filters"""
        try:
            # Test basic transaction retrieval
            response = self.session.get(f"{API_BASE}/transactions")
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    transactions_data = data['data']
                    transactions = transactions_data.get('transactions', [])
                    total = transactions_data.get('total', 0)
                    
                    self.log_result(
                        "GET /api/transactions", 
                        True, 
                        f"Retrieved {len(transactions)} transactions (total: {total})",
                        {'transaction_count': len(transactions), 'total': total}
                    )
                    
                    # Test with pagination
                    response_page2 = self.session.get(f"{API_BASE}/transactions?page=2&limit=5")
                    if response_page2.status_code == 200:
                        page2_data = response_page2.json()
                        if page2_data.get('success'):
                            self.log_result(
                                "GET /api/transactions (pagination)", 
                                True, 
                                "Pagination working correctly"
                            )
                    
                    # Test with search filter
                    response_search = self.session.get(f"{API_BASE}/transactions?search=salary")
                    if response_search.status_code == 200:
                        search_data = response_search.json()
                        if search_data.get('success'):
                            search_results = search_data['data']['transactions']
                            self.log_result(
                                "GET /api/transactions (search)", 
                                True, 
                                f"Search returned {len(search_results)} results"
                            )
                    
                    return transactions
                else:
                    self.log_result("GET /api/transactions", False, "Invalid response format", data)
            else:
                self.log_result("GET /api/transactions", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("GET /api/transactions", False, f"Request failed: {str(e)}")
        return []
    
    def test_create_transaction(self, account_id=None, category_id=None):
        """Test POST /api/transactions endpoint"""
        try:
            # Use provided IDs or get from existing data
            if not account_id:
                accounts = self.test_get_accounts()
                if not accounts:
                    self.log_result("POST /api/transactions", False, "No accounts available for testing")
                    return None
                account_id = accounts[0]['id']
            
            if not category_id:
                categories = self.test_get_categories()
                if not categories:
                    self.log_result("POST /api/transactions", False, "No categories available for testing")
                    return None
                category_id = categories[0]['id']
            
            test_transaction = {
                "amount": 75.25,
                "description": f"Test Transaction {uuid.uuid4().hex[:8]}",
                "date": datetime.now().isoformat(),
                "accountId": account_id,
                "categoryId": category_id
            }
            
            response = self.session.post(
                f"{API_BASE}/transactions",
                json=test_transaction,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    transaction = data['data']
                    self.created_resources['transactions'].append(transaction['id'])
                    self.log_result(
                        "POST /api/transactions", 
                        True, 
                        f"Created transaction '{transaction['description']}'",
                        {'transaction_id': transaction['id'], 'amount': transaction['amount']}
                    )
                    return transaction
                else:
                    self.log_result("POST /api/transactions", False, "Invalid response format", data)
            else:
                self.log_result("POST /api/transactions", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("POST /api/transactions", False, f"Request failed: {str(e)}")
        return None
    
    def test_analytics(self):
        """Test GET /api/analytics endpoint"""
        try:
            response = self.session.get(f"{API_BASE}/analytics")
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    analytics = data['data']
                    required_fields = ['totalIncome', 'totalExpense', 'netSavings', 'transactionCount']
                    
                    missing_fields = [field for field in required_fields if field not in analytics]
                    if not missing_fields:
                        self.log_result(
                            "GET /api/analytics", 
                            True, 
                            f"Analytics retrieved successfully",
                            {
                                'total_income': analytics['totalIncome'],
                                'total_expense': analytics['totalExpense'],
                                'net_savings': analytics['netSavings'],
                                'transaction_count': analytics['transactionCount']
                            }
                        )
                        
                        # Test with date filters
                        start_date = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                        end_date = datetime.now().strftime('%Y-%m-%d')
                        
                        response_filtered = self.session.get(
                            f"{API_BASE}/analytics?startDate={start_date}&endDate={end_date}"
                        )
                        
                        if response_filtered.status_code == 200:
                            filtered_data = response_filtered.json()
                            if filtered_data.get('success'):
                                self.log_result(
                                    "GET /api/analytics (date filter)", 
                                    True, 
                                    "Date filtering working correctly"
                                )
                        
                        return analytics
                    else:
                        self.log_result("GET /api/analytics", False, f"Missing fields: {missing_fields}", analytics)
                else:
                    self.log_result("GET /api/analytics", False, "Invalid response format", data)
            else:
                self.log_result("GET /api/analytics", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("GET /api/analytics", False, f"Request failed: {str(e)}")
        return None
    
    def test_upload_endpoint(self):
        """Test POST /api/upload endpoint"""
        try:
            # Create a simple CSV content for testing
            csv_content = """Date,Description,Amount
2024-06-01,Test Income,1000
2024-06-02,Test Expense,-50
2024-06-03,Another Transaction,200"""
            
            # Create a temporary file-like object
            files = {
                'file': ('test_transactions.csv', csv_content, 'text/csv')
            }
            
            response = self.session.post(f"{API_BASE}/upload", files=files)
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    upload_data = data['data']
                    required_fields = ['preview', 'columns', 'totalRows']
                    
                    missing_fields = [field for field in required_fields if field not in upload_data]
                    if not missing_fields:
                        self.log_result(
                            "POST /api/upload", 
                            True, 
                            f"File uploaded and parsed successfully",
                            {
                                'total_rows': upload_data['totalRows'],
                                'columns': upload_data['columns']
                            }
                        )
                        return upload_data
                    else:
                        self.log_result("POST /api/upload", False, f"Missing fields: {missing_fields}", upload_data)
                else:
                    self.log_result("POST /api/upload", False, "Invalid response format", data)
            else:
                self.log_result("POST /api/upload", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("POST /api/upload", False, f"Request failed: {str(e)}")
        return None
    
    def test_import_transactions(self):
        """Test POST /api/import endpoint"""
        try:
            # First get accounts and categories for import
            accounts = self.test_get_accounts()
            categories = self.test_get_categories()
            
            if not accounts or not categories:
                self.log_result("POST /api/import", False, "No accounts or categories available for import testing")
                return None
            
            # Sample import data
            import_data = {
                "data": [
                    {"Date": "2024-06-01", "Description": "Import Test 1", "Amount": "100"},
                    {"Date": "2024-06-02", "Description": "Import Test 2", "Amount": "-50"}
                ],
                "mapping": {
                    "date": "Date",
                    "description": "Description",
                    "amount": "Amount"
                },
                "defaultAccountId": accounts[0]['id'],
                "defaultCategoryId": categories[0]['id']
            }
            
            response = self.session.post(
                f"{API_BASE}/import",
                json=import_data,
                headers={'Content-Type': 'application/json'}
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success') and 'data' in data:
                    import_result = data['data']
                    self.log_result(
                        "POST /api/import", 
                        True, 
                        f"Imported {import_result.get('imported', 0)} transactions",
                        import_result
                    )
                    return import_result
                else:
                    self.log_result("POST /api/import", False, "Invalid response format", data)
            else:
                self.log_result("POST /api/import", False, f"HTTP {response.status_code}", response.text)
        except Exception as e:
            self.log_result("POST /api/import", False, f"Request failed: {str(e)}")
        return None
    
    def test_account_balance_update(self):
        """Test that account balance updates correctly when transactions are created"""
        try:
            # Create a test account
            test_account = self.test_create_account()
            if not test_account:
                self.log_result("Account Balance Update", False, "Could not create test account")
                return
            
            initial_balance = test_account['balance']
            
            # Get a category for the transaction
            categories = self.test_get_categories()
            if not categories:
                self.log_result("Account Balance Update", False, "No categories available")
                return
            
            # Find an expense category
            expense_category = next((cat for cat in categories if cat['type'] == 'EXPENSE'), categories[0])
            
            # Create a transaction
            transaction_amount = 100.0
            transaction = self.test_create_transaction(test_account['id'], expense_category['id'])
            
            if transaction:
                # Check if account balance was updated
                updated_accounts = self.test_get_accounts()
                updated_account = next((acc for acc in updated_accounts if acc['id'] == test_account['id']), None)
                
                if updated_account:
                    expected_balance = initial_balance - transaction_amount  # Expense reduces balance
                    actual_balance = updated_account['balance']
                    
                    if abs(actual_balance - expected_balance) < 0.01:  # Allow for floating point precision
                        self.log_result(
                            "Account Balance Update", 
                            True, 
                            f"Balance updated correctly: {initial_balance} -> {actual_balance}",
                            {
                                'initial_balance': initial_balance,
                                'transaction_amount': transaction_amount,
                                'expected_balance': expected_balance,
                                'actual_balance': actual_balance
                            }
                        )
                    else:
                        self.log_result(
                            "Account Balance Update", 
                            False, 
                            f"Balance not updated correctly. Expected: {expected_balance}, Actual: {actual_balance}"
                        )
                else:
                    self.log_result("Account Balance Update", False, "Could not find updated account")
            else:
                self.log_result("Account Balance Update", False, "Could not create test transaction")
        except Exception as e:
            self.log_result("Account Balance Update", False, f"Test failed: {str(e)}")
    
    def test_export_csv(self):
        """Test CSV export functionality - GET /api/export?format=csv"""
        try:
            response = self.session.get(f"{API_BASE}/export?format=csv", timeout=30)
            
            if response.status_code == 200:
                # Verify Content-Type
                expected_content_type = "text/csv"
                actual_content_type = response.headers.get('Content-Type', '')
                
                if expected_content_type in actual_content_type:
                    # Verify Content-Disposition header
                    content_disposition = response.headers.get('Content-Disposition', '')
                    if 'attachment' in content_disposition and '.csv' in content_disposition:
                        # Verify CSV content structure
                        csv_content = response.text
                        lines = csv_content.strip().split('\n')
                        
                        if len(lines) > 0:
                            headers = lines[0].split(',')
                            expected_headers = ['Date', 'Description', 'Amount', 'Category', 'Subcategory', 'Account', 'Type']
                            
                            # Check if all expected headers are present
                            headers_match = all(header in headers for header in expected_headers)
                            if headers_match:
                                # Test CSV parsing
                                try:
                                    import csv
                                    from io import StringIO
                                    csv_reader = csv.DictReader(StringIO(csv_content))
                                    rows = list(csv_reader)
                                    
                                    self.log_result(
                                        "GET /api/export (CSV)", 
                                        True, 
                                        f"CSV export successful with {len(rows)} rows",
                                        {
                                            'headers': headers,
                                            'row_count': len(rows),
                                            'content_type': actual_content_type,
                                            'content_disposition': content_disposition
                                        }
                                    )
                                    return True
                                except Exception as parse_error:
                                    self.log_result("GET /api/export (CSV)", False, f"CSV parsing error: {parse_error}")
                            else:
                                self.log_result("GET /api/export (CSV)", False, f"CSV headers mismatch. Expected: {expected_headers}, Got: {headers}")
                        else:
                            self.log_result("GET /api/export (CSV)", False, "CSV content is empty")
                    else:
                        self.log_result("GET /api/export (CSV)", False, f"Invalid Content-Disposition header: {content_disposition}")
                else:
                    self.log_result("GET /api/export (CSV)", False, f"Invalid Content-Type. Expected: {expected_content_type}, Got: {actual_content_type}")
            else:
                self.log_result("GET /api/export (CSV)", False, f"HTTP {response.status_code}: {response.text}")
        except Exception as e:
            self.log_result("GET /api/export (CSV)", False, f"Request failed: {str(e)}")
        return False
    
    def test_export_excel(self):
        """Test Excel export functionality - GET /api/export?format=xlsx"""
        try:
            response = self.session.get(f"{API_BASE}/export?format=xlsx", timeout=30)
            
            if response.status_code == 200:
                # Verify Content-Type
                expected_content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                actual_content_type = response.headers.get('Content-Type', '')
                
                if expected_content_type in actual_content_type:
                    # Verify Content-Disposition header
                    content_disposition = response.headers.get('Content-Disposition', '')
                    if 'attachment' in content_disposition and '.xlsx' in content_disposition:
                        # Verify Excel file format
                        try:
                            from io import BytesIO
                            import openpyxl
                            
                            excel_data = BytesIO(response.content)
                            workbook = openpyxl.load_workbook(excel_data)
                            
                            if 'Transactions' in workbook.sheetnames:
                                worksheet = workbook['Transactions']
                                
                                # Get headers from first row
                                headers = []
                                for cell in worksheet[1]:
                                    if cell.value:
                                        headers.append(cell.value)
                                
                                expected_headers = ['Date', 'Description', 'Amount', 'Category', 'Subcategory', 'Account', 'Type']
                                headers_match = all(header in headers for header in expected_headers)
                                
                                if headers_match:
                                    data_rows = worksheet.max_row - 1  # Subtract header row
                                    
                                    self.log_result(
                                        "GET /api/export (Excel)", 
                                        True, 
                                        f"Excel export successful with {data_rows} rows",
                                        {
                                            'headers': headers,
                                            'row_count': data_rows,
                                            'content_type': actual_content_type,
                                            'content_disposition': content_disposition,
                                            'worksheets': workbook.sheetnames
                                        }
                                    )
                                    workbook.close()
                                    return True
                                else:
                                    self.log_result("GET /api/export (Excel)", False, f"Excel headers mismatch. Expected: {expected_headers}, Got: {headers}")
                            else:
                                self.log_result("GET /api/export (Excel)", False, "'Transactions' worksheet not found")
                            
                            workbook.close()
                        except Exception as excel_error:
                            self.log_result("GET /api/export (Excel)", False, f"Excel file validation error: {excel_error}")
                    else:
                        self.log_result("GET /api/export (Excel)", False, f"Invalid Content-Disposition header: {content_disposition}")
                else:
                    self.log_result("GET /api/export (Excel)", False, f"Invalid Content-Type. Expected: {expected_content_type}, Got: {actual_content_type}")
            else:
                self.log_result("GET /api/export (Excel)", False, f"HTTP {response.status_code}: {response.text}")
        except Exception as e:
            self.log_result("GET /api/export (Excel)", False, f"Request failed: {str(e)}")
        return False
    
    def test_export_default_format(self):
        """Test default export format (should be CSV) - GET /api/export"""
        try:
            response = self.session.get(f"{API_BASE}/export", timeout=30)
            
            if response.status_code == 200:
                # Should default to CSV
                expected_content_type = "text/csv"
                actual_content_type = response.headers.get('Content-Type', '')
                
                if expected_content_type in actual_content_type:
                    # Verify it's the same as explicit CSV format
                    csv_response = self.session.get(f"{API_BASE}/export?format=csv", timeout=30)
                    if csv_response.status_code == 200:
                        if response.text == csv_response.text:
                            self.log_result(
                                "GET /api/export (Default Format)", 
                                True, 
                                "Default format correctly defaults to CSV",
                                {'content_type': actual_content_type}
                            )
                            return True
                        else:
                            self.log_result("GET /api/export (Default Format)", False, "Default format differs from explicit CSV format")
                    else:
                        self.log_result("GET /api/export (Default Format)", False, "Could not compare with explicit CSV format")
                else:
                    self.log_result("GET /api/export (Default Format)", False, f"Default format not CSV. Got: {actual_content_type}")
            else:
                self.log_result("GET /api/export (Default Format)", False, f"HTTP {response.status_code}: {response.text}")
        except Exception as e:
            self.log_result("GET /api/export (Default Format)", False, f"Request failed: {str(e)}")
        return False
    
    def test_export_data_integrity(self):
        """Test that export includes all expected transaction data"""
        try:
            # First, get transactions via the regular API to compare
            transactions_response = self.session.get(f"{API_BASE}/transactions?limit=1000", timeout=30)
            
            if transactions_response.status_code != 200:
                self.log_result("Export Data Integrity", False, "Could not fetch transactions for comparison")
                return False
                
            transactions_data = transactions_response.json()
            if not transactions_data.get('success'):
                self.log_result("Export Data Integrity", False, "Transactions API returned unsuccessful response")
                return False
                
            api_transactions = transactions_data.get('data', {}).get('transactions', [])
            
            # Get CSV export for comparison
            csv_response = self.session.get(f"{API_BASE}/export?format=csv", timeout=30)
            
            if csv_response.status_code == 200:
                csv_content = csv_response.text
                csv_lines = csv_content.strip().split('\n')
                csv_data_rows = len(csv_lines) - 1  # Subtract header
                
                # The export should include all transactions (not paginated)
                if csv_data_rows >= len(api_transactions):
                    # Check if CSV contains the expected fields
                    if len(csv_lines) > 1:
                        import csv
                        from io import StringIO
                        csv_reader = csv.DictReader(StringIO(csv_content))
                        first_csv_row = next(csv_reader)
                        
                        # Check if key fields are present
                        required_fields = ['Date', 'Description', 'Amount', 'Category', 'Account', 'Type']
                        missing_fields = [field for field in required_fields if not first_csv_row.get(field)]
                        
                        if not missing_fields:
                            self.log_result(
                                "Export Data Integrity", 
                                True, 
                                f"Export contains all required fields and {csv_data_rows} rows",
                                {
                                    'api_transactions': len(api_transactions),
                                    'export_rows': csv_data_rows,
                                    'required_fields': required_fields
                                }
                            )
                            return True
                        else:
                            self.log_result("Export Data Integrity", False, f"Missing fields in export: {missing_fields}")
                    else:
                        self.log_result("Export Data Integrity", False, "Export has no data rows")
                else:
                    self.log_result("Export Data Integrity", False, f"Export has fewer rows ({csv_data_rows}) than API transactions ({len(api_transactions)})")
            else:
                self.log_result("Export Data Integrity", False, f"Could not get CSV export: {csv_response.status_code}")
        except Exception as e:
            self.log_result("Export Data Integrity", False, f"Test failed: {str(e)}")
        return False
    
    def run_all_tests(self):
        """Run all backend API tests"""
        print(f"\n🚀 Starting Personal Finance Dashboard Backend API Tests")
        print(f"📍 Base URL: {API_BASE}")
        print("=" * 80)
        
        # Test all endpoints
        self.test_get_accounts()
        self.test_create_account()
        self.test_get_categories()
        self.test_create_category()
        self.test_get_transactions()
        self.test_create_transaction()
        self.test_analytics()
        self.test_upload_endpoint()
        self.test_import_transactions()
        self.test_account_balance_update()
        
        # Test export functionality
        print("\n📤 Testing Export Functionality")
        print("-" * 40)
        self.test_export_csv()
        self.test_export_excel()
        self.test_export_default_format()
        self.test_export_data_integrity()
        
        # Summary
        print("\n" + "=" * 80)
        print("📊 TEST SUMMARY")
        print("=" * 80)
        
        passed = sum(1 for result in self.test_results if result['success'])
        failed = len(self.test_results) - passed
        
        print(f"✅ Passed: {passed}")
        print(f"❌ Failed: {failed}")
        print(f"📈 Success Rate: {(passed/len(self.test_results)*100):.1f}%")
        
        if failed > 0:
            print("\n🔍 FAILED TESTS:")
            for result in self.test_results:
                if not result['success']:
                    print(f"   ❌ {result['test']}: {result['message']}")
        
        print("\n" + "=" * 80)
        return passed, failed

if __name__ == "__main__":
    tester = FinanceAPITester()
    passed, failed = tester.run_all_tests()
    
    # Exit with appropriate code
    exit(0 if failed == 0 else 1)