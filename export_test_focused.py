#!/usr/bin/env python3
"""
Focused Export Functionality Test for Personal Finance Dashboard
Tests only the /api/export endpoint with CSV and Excel formats
"""

import requests
import json
import os
from io import BytesIO, StringIO
import csv
import openpyxl

# Get base URL from environment
BASE_URL = "https://finview-13.preview.emergentagent.com"
API_BASE = f"{BASE_URL}/api"

def test_export_functionality():
    """Run focused export functionality tests"""
    print("🚀 Testing Personal Finance Dashboard Export Functionality")
    print(f"📍 API Base: {API_BASE}")
    print("=" * 70)
    
    results = []
    
    # Test 1: CSV Export
    print("\n1️⃣ Testing CSV Export (GET /api/export?format=csv)")
    try:
        response = requests.get(f"{API_BASE}/export?format=csv", timeout=30)
        
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type', '')
            content_disposition = response.headers.get('Content-Disposition', '')
            
            print(f"   ✅ Status: {response.status_code}")
            print(f"   ✅ Content-Type: {content_type}")
            print(f"   ✅ Content-Disposition: {content_disposition}")
            
            # Verify CSV structure
            csv_content = response.text
            lines = csv_content.strip().split('\n')
            
            if len(lines) > 0:
                headers = lines[0].split(',')
                expected_headers = ['Date', 'Description', 'Amount', 'Category', 'Subcategory', 'Account', 'Type']
                
                print(f"   ✅ Headers: {headers}")
                print(f"   ✅ Data rows: {len(lines) - 1}")
                
                # Test CSV parsing
                csv_reader = csv.DictReader(StringIO(csv_content))
                rows = list(csv_reader)
                print(f"   ✅ Parsed rows: {len(rows)}")
                
                if len(rows) > 0:
                    sample_row = rows[0]
                    print(f"   ✅ Sample row: {dict(sample_row)}")
                
                results.append(("CSV Export", True, f"Successfully exported {len(rows)} transactions"))
            else:
                results.append(("CSV Export", False, "Empty CSV content"))
        else:
            results.append(("CSV Export", False, f"HTTP {response.status_code}: {response.text}"))
    except Exception as e:
        results.append(("CSV Export", False, f"Error: {e}"))
    
    # Test 2: Excel Export
    print("\n2️⃣ Testing Excel Export (GET /api/export?format=xlsx)")
    try:
        response = requests.get(f"{API_BASE}/export?format=xlsx", timeout=30)
        
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type', '')
            content_disposition = response.headers.get('Content-Disposition', '')
            
            print(f"   ✅ Status: {response.status_code}")
            print(f"   ✅ Content-Type: {content_type}")
            print(f"   ✅ Content-Disposition: {content_disposition}")
            
            # Verify Excel structure
            excel_data = BytesIO(response.content)
            workbook = openpyxl.load_workbook(excel_data)
            
            print(f"   ✅ Worksheets: {workbook.sheetnames}")
            
            if 'Transactions' in workbook.sheetnames:
                worksheet = workbook['Transactions']
                
                # Get headers
                headers = []
                for cell in worksheet[1]:
                    if cell.value:
                        headers.append(cell.value)
                
                data_rows = worksheet.max_row - 1
                print(f"   ✅ Headers: {headers}")
                print(f"   ✅ Data rows: {data_rows}")
                
                if data_rows > 0:
                    # Sample first data row
                    sample_row = []
                    for cell in worksheet[2]:
                        sample_row.append(cell.value)
                    print(f"   ✅ Sample row: {sample_row}")
                
                results.append(("Excel Export", True, f"Successfully exported {data_rows} transactions"))
            else:
                results.append(("Excel Export", False, "Transactions worksheet not found"))
            
            workbook.close()
        else:
            results.append(("Excel Export", False, f"HTTP {response.status_code}: {response.text}"))
    except Exception as e:
        results.append(("Excel Export", False, f"Error: {e}"))
    
    # Test 3: Default Format
    print("\n3️⃣ Testing Default Format (GET /api/export)")
    try:
        response = requests.get(f"{API_BASE}/export", timeout=30)
        
        if response.status_code == 200:
            content_type = response.headers.get('Content-Type', '')
            print(f"   ✅ Status: {response.status_code}")
            print(f"   ✅ Content-Type: {content_type}")
            
            if 'text/csv' in content_type:
                # Compare with explicit CSV
                csv_response = requests.get(f"{API_BASE}/export?format=csv", timeout=30)
                if csv_response.status_code == 200 and response.text == csv_response.text:
                    results.append(("Default Format", True, "Correctly defaults to CSV"))
                    print("   ✅ Matches explicit CSV format")
                else:
                    results.append(("Default Format", False, "Does not match explicit CSV"))
            else:
                results.append(("Default Format", False, f"Not CSV format: {content_type}"))
        else:
            results.append(("Default Format", False, f"HTTP {response.status_code}: {response.text}"))
    except Exception as e:
        results.append(("Default Format", False, f"Error: {e}"))
    
    # Test 4: Data Integrity
    print("\n4️⃣ Testing Data Integrity")
    try:
        # Get transactions from API
        transactions_response = requests.get(f"{API_BASE}/transactions?limit=1000", timeout=30)
        
        if transactions_response.status_code == 200:
            transactions_data = transactions_response.json()
            if transactions_data.get('success'):
                api_transactions = transactions_data.get('data', {}).get('transactions', [])
                print(f"   ✅ API transactions: {len(api_transactions)}")
                
                # Get CSV export
                csv_response = requests.get(f"{API_BASE}/export?format=csv", timeout=30)
                if csv_response.status_code == 200:
                    csv_content = csv_response.text
                    csv_lines = csv_content.strip().split('\n')
                    csv_data_rows = len(csv_lines) - 1
                    
                    print(f"   ✅ Export rows: {csv_data_rows}")
                    
                    # Check required fields
                    csv_reader = csv.DictReader(StringIO(csv_content))
                    first_row = next(csv_reader)
                    required_fields = ['Date', 'Description', 'Amount', 'Category', 'Account', 'Type']
                    missing_fields = [field for field in required_fields if not first_row.get(field)]
                    
                    if not missing_fields and csv_data_rows >= len(api_transactions):
                        results.append(("Data Integrity", True, f"All fields present, {csv_data_rows} rows exported"))
                        print(f"   ✅ All required fields present: {required_fields}")
                    else:
                        results.append(("Data Integrity", False, f"Missing fields: {missing_fields} or insufficient rows"))
                else:
                    results.append(("Data Integrity", False, "Could not get CSV for comparison"))
            else:
                results.append(("Data Integrity", False, "Transactions API unsuccessful"))
        else:
            results.append(("Data Integrity", False, f"Could not get transactions: {transactions_response.status_code}"))
    except Exception as e:
        results.append(("Data Integrity", False, f"Error: {e}"))
    
    # Summary
    print("\n" + "=" * 70)
    print("📊 EXPORT FUNCTIONALITY TEST SUMMARY")
    print("=" * 70)
    
    passed = sum(1 for _, success, _ in results if success)
    total = len(results)
    
    for test_name, success, message in results:
        status = "✅ PASS" if success else "❌ FAIL"
        print(f"{status}: {test_name} - {message}")
    
    print(f"\n🎯 Results: {passed}/{total} tests passed ({(passed/total*100):.1f}%)")
    
    if passed == total:
        print("🎉 All export functionality tests PASSED!")
        return True
    else:
        print("⚠️ Some export functionality tests FAILED!")
        return False

if __name__ == "__main__":
    success = test_export_functionality()
    exit(0 if success else 1)